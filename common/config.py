# -*- coding: utf-8 -*-
"""
调取配置文件和屏幕分辨率的代码
"""
import os
import sys
import json
import re
import openpyxl
import logging
from common.adb import ADB
from common import constant

log = logging.getLogger('FTX.config')

class LoadError(Exception):
    def __init__(self, file_name):
        msg = "{} 不存在".format(file_name)
        super(LoadError, self).__init__(msg)

class DuplicateKey(Exception):
    def __init__(self, key, file_name):
        msg = "{} 的'{}'重复".format(file_name, key)
        super(LoadError, self).__init__(msg)
    
class OpErrorType(Exception):
    def __init__(self, key):
        msg = "操作'{}'不支持".format(key)
        super(OpErrorType, self).__init__(msg)

class OpErrorAttr(Exception):
    def __init__(self, i, file_name):
        msg = "{}的第{}行配置有误".format(file_name, i)
        super(OpErrorAttr, self).__init__(msg)
    
class Component(object):
    
    def __init__(self):
        super(Component, self).__init__()
        self.name = "组件.xlsx"
        cpts = None
    
    def load(self):
        cpts = {}
        xlsx = os.path.join(constant.CONFIG_PATH, self.name)
        if not os.path.isfile(xlsx):
            raise LoadError(self.name)
        wb = openpyxl.load_workbook(xlsx)
        ws = wb.active  # 当前活跃的表单
        try:
            name = None
            for i, row in enumerate(ws.iter_rows(min_row=3)):
                _name = row[0].value
                if _name:
                    if _name not in constant.STEPS:
                        raise OpErrorType(_name)
                    if _name in cpts:
                        raise DuplicateKey(cpt['name'], self.name)
                    name = _name
                    cpts[name] = []
                cpt = self._parse_row(i, row)
                cpts[name].append(cpt)
        except Exception as ex:
            log.exception(str(ex))
            self.cpts = None
            return None
        self.cpts = cpts
        return cpts
        
    def _parse_row(self, i, row):
        op = row[1].value
        if op not in constant.COMPONENTS:
            raise OpErrorType(op)
        _type = constant.COMPONENTS[op]
        cpt = {
            "type": _type,
            "op": op,
            "start": (row[2].value, row[3].value),
            "end": (row[4].value, row[5].value),
            "txt": row[6].value,
            "sleep": row[7].value,
            "desc": row[8].value
        }
        if not self._is_row_ok(cpt):
            raise OpErrorAttr(i, self.name)
        return cpt
    
    def _is_row_ok(self, cpt):
        if cpt['type'] in ['tap', 'username', 'password']:
            return self._chk_tap(cpt['start'])
        if cpt['type'] == 'swipe':
            return self._chk_swipe(cpt['start'], cpt['end'])
        if cpt['type'] == 'input':
            return self._chk_input(cpt['start'], cpt['txt'])
        return False

    def _chk_tap(self, pos):
        x, y = pos
        if not x or not y:
            return False
        return True
    
    def _chk_swipe(self, start, end):
        if not self._chk_tap(start):
            return False
        if not self._chk_tap(end):
            return False
        return True
    
    def _chk_input(self, pos, txt):
        if not self._chk_tap(start):
            return False
        return False if not txt else True

class Step(object):
    
    def __init__(self):
        super(Step, self).__init__()
        self.name = "步骤.xlsx"
        self.steps = None
    
    def load(self):
        steps = []
        xlsx = os.path.join(constant.CONFIG_PATH, self.name)
        if not os.path.isfile(xlsx):
            raise LoadError(self.name)
        wb = openpyxl.load_workbook(xlsx)
        ws = wb.active  # 当前活跃的表单
        try:
            for i, row in enumerate(ws.iter_rows(min_row=2)):
                steps.append(self._parse_row(i, row))
        except Exception as ex:
            log.exception(str(ex))
            self.steps = None
            return None
        self.steps = steps
        return steps
        
    def _parse_row(self, i, row):
        op = row[0].value
        if op not in constant.STEPS:
            raise OpErrorType(op)
        _type = constant.STEPS[op]
        cpt = {
            "name": row[0].value,
            "type": _type,
            "username": row[1].value,
            "password": row[2].value
        }
        if not self._is_row_ok(cpt):
            raise OpErrorAttr(i, self.name)
        return cpt
    
    def _is_row_ok(self, cpt):
        if cpt['type'] == 'login':
            return self._chk_login(cpt['username'], cpt['password'])
        return False
    
    def _chk_login(self, username, password):
        if not username or not password:
            return False
        return True
